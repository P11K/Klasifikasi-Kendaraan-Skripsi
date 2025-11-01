import cv2
import torch
import threading
import tkinter as tk
from tkinter import ttk, filedialog, simpledialog, messagebox
from PIL import Image, ImageTk
from ultralytics import YOLO
from queue import Queue
import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl import Workbook
import os
from deep_sort_realtime.deepsort_tracker import DeepSort
import math
import time
# from collections import deque


class VehicleCounterApp:
        # Load DeepSort tracker
    tracker = DeepSort(max_age=5)

    # Simpan lintasan kendaraan
    trajectory = {}
    
    def __init__(self, root):
        self.roi_dict = {} 
        self.root = root
        self.root.title("Vehicle Counting System 🚦")
        self.root.state('zoomed')
        self.roi_colors = ["#FF0000", "#00FF00", "#0000FF"] 
        self.device = 'cuda' if torch.cuda.is_available() else 'cpu'
        self.model = YOLO('runs/detect/fine/weights/best.pt').to(self.device)
        # self.model = YOLO("yolov8x.pt").to(self.device)  # Ganti sesuai path model Anda
        
        self.vehicle_classes = {
            0: ("Bus", "#4CAF50"),
            1: ("Motor", "#2196F3"),
            2: ("Truck", "#FFC107"),
            3: ("Mobil", "#E91E63")
        }
        # self.vehicle_classes = {
        #     5: ("Bus", "#4CAF50"),
        #     3: ("Motor", "#2196F3"),
        #     7: ("Truck", "#FFC107"),
        #     2: ("Mobil", "#E91E63")
        # }

        self.track_history = {}
        self.lines_info = []
        self.frame_queue = Queue(maxsize=10)
        self.lock = threading.Lock()
        self.running = True
        self.points = []
        self.processing_started = False
        self.total_vehicle_count = {name: 0 for name, _ in self.vehicle_classes.values()}

        self.setup_gui()

    def setup_gui(self):
        style = ttk.Style()
        style.configure("TFrame", background="#f0f2f5")
        style.configure("TLabel", background="#f0f2f5", font=('Segoe UI', 11))
        style.configure("Header.TLabel", font=('Segoe UI', 16, 'bold'), background="#f0f2f5")
        style.configure("Road.TLabelframe", background="#ffffff", relief="solid", borderwidth=1)
        style.configure("Road.TLabelframe.Label", font=('Segoe UI', 13, 'bold'))

        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        self.counter_frame = ttk.LabelFrame(self.main_frame, text="Statistik Kendaraan", style="Road.TLabelframe")
        self.counter_frame.pack(side=tk.LEFT, fill=tk.Y, padx=15, pady=15)

        self.save_btn = ttk.Button(self.counter_frame, text="💾 Simpan ke Excel", command=self.save_to_excel, state='disabled')
        self.save_btn.pack(side=tk.BOTTOM, pady=(5, 15))


        self.exit_btn = ttk.Button(self.counter_frame, text="❌ Keluar", command=self.shutdown)
        self.exit_btn.pack(side=tk.BOTTOM, pady=(5, 15))

        self.upload_btn = ttk.Button(self.counter_frame, text="🎥 Upload Video", command=self.upload_video)
        self.upload_btn.pack(side=tk.BOTTOM, pady=(5, 15))

        self.total_count_label = ttk.Label(self.counter_frame, text="Total Semua Kendaraan: 0", style="Header.TLabel")
        self.total_count_label.pack(side=tk.TOP, pady=(0, 20))

        self.counter_labels = {}
        self.road_frames = {}
        
        video_name = "Belum ada video"
        self.video_frame = ttk.LabelFrame(self.main_frame, text=video_name, style="Road.TLabelframe")
        self.video_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=15, pady=15)

        self.canvas = tk.Canvas(self.video_frame, bg="black")
        self.canvas.pack(fill=tk.BOTH, expand=True)
        self.canvas.bind("<Button-1>", self.add_point)

    def reset_state(self):
        self.points.clear()
        self.lines_info.clear()
        self.track_history.clear()
        self.total_vehicle_count = {name: 0 for name, _ in self.vehicle_classes.values()}

        for widget in self.counter_frame.winfo_children():
            if widget not in (self.upload_btn, self.exit_btn, self.total_count_label, self.save_btn):
                widget.destroy()

        self.counter_labels.clear()
        self.road_frames.clear()

        self.canvas.delete("all")
        self.processing_started = False

    def upload_video(self):
        video_file = filedialog.askopenfilename(filetypes=[("Video Files", "*.mp4;, *.avi; *.mov")], title="Pilih Video")
        if video_file != "*.mp4;":
            if hasattr(self, 'cap') and self.cap.isOpened():
                self.cap.release()

            self.reset_state()
            self.save_btn.config(state='disabled')
            self.model = YOLO('runs/detect/fine/weights/best.pt').to(self.device)
            # self.model = YOLO("yolov8x.pt").to(self.device)  # Ganti sesuai path model Anda

            self.video_path = video_file
            self.video_frame.config(text=os.path.basename(self.video_path))
            self.cap = cv2.VideoCapture(self.video_path)

            ret, frame = self.cap.read()
            if not ret:
                raise ValueError("Tidak bisa membaca frame video")

            self.original_width = frame.shape[1]
            self.original_height = frame.shape[0]

            self.canvas_width = self.canvas.winfo_width()
            self.canvas_height = self.canvas.winfo_height()

            aspect_ratio = self.original_width / self.original_height
            if self.canvas_width / self.canvas_height > aspect_ratio:
                new_width = int(self.canvas_height * aspect_ratio)
                new_height = self.canvas_height
            else:
                new_width = self.canvas_width
                new_height = int(self.canvas_width / aspect_ratio)

            frame_resized = cv2.resize(frame, (new_width, new_height))
            frame_rgb = cv2.cvtColor(frame_resized, cv2.COLOR_BGR2RGB)
            image = Image.fromarray(frame_rgb)
            self.tk_img = ImageTk.PhotoImage(image=image)

            self.canvas.create_image(0, 0, anchor=tk.NW, image=self.tk_img)
            self.canvas.config(width=new_width, height=new_height)

            self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)

            self.frame_width = new_width
            self.frame_height = new_height
        else:
            messagebox.showerror("Error","Format Tidak Didukung")
    def add_point(self, event):
        x, y = event.x, event.y
        self.points.append((x, y))
        roi_colors = ["#FF0000", "#00FF00", "#0000FF"] 
        current_roi_index = len(self.lines_info)
        color = roi_colors[current_roi_index % 3]  # Cycle 3 warna
        # Gambar titik dengan warna yang sesuai
        self.canvas.create_oval(x-3, y-3, x+3, y+3, fill=color)
        if len(self.points) % 2 == 0:
            x1, y1 = self.points[-2]
            x2, y2 = self.points[-1]
            self.canvas.create_line(x1, y1, x2, y2, fill=color, width=2)

            road_name = simpledialog.askstring("Input", "Masukkan nama jalan untuk ROI ini:")
            allowed_directions_str = simpledialog.askstring("Arah Diizinkan", "Masukkan arah yang diizinkan (pisahkan dengan koma):\nContoh: atas,bawah")
            allowed_directions = [a.strip().lower() for a in allowed_directions_str.split(',')] if allowed_directions_str else []

            if road_name:
                counter = {name: 0 for name, _ in self.vehicle_classes.values()}
                self.lines_info.append({
                    'points': ((x1, y1), (x2, y2)),
                    'name': road_name,
                    'counter': counter,
                    'allowed_directions': allowed_directions if allowed_directions else [],
                    'color': color
                })
                
                self.roi_dict[road_name] = ((x1, y1), (x2, y2))  # Simpan ROI ke dalam roi_dict
                road_frame = ttk.LabelFrame(self.counter_frame, text=road_name, style="Road.TLabelframe")
                road_frame.pack(anchor="w", padx=10, pady=10, fill=tk.X)
                self.road_frames[road_name] = road_frame

                for veh_name, color in self.vehicle_classes.values():
                    lbl = ttk.Label(road_frame, text=f"{veh_name}: 0", foreground=color)
                    lbl.pack(anchor="w", padx=10, pady=2)
                    self.counter_labels[(road_name, veh_name)] = lbl

                total_lbl = ttk.Label(road_frame, text=f"Total: 0", font=('Segoe UI', 11, 'bold'))
                total_lbl.pack(anchor="w", pady=(5, 0))
                self.counter_labels[(road_name, 'Total')] = total_lbl

            if len(self.lines_info) >= 3 and not self.processing_started:
                self.processing_thread = threading.Thread(target=self.process_video)
                self.processing_thread.start()
                self.processing_started = True

    def process_video(self):
        prev_time = 0
        fps_list = []
        while self.running:
            ret, frame = self.cap.read()
            if not ret:
                break
            
            current_time = time.time()
            fps = 1 / (current_time - prev_time) if prev_time != 0 else 0
            prev_time = current_time
            fps_list.append(fps)
            
            frame_resized = cv2.resize(frame, (self.frame_width, self.frame_height))
            cv2.putText(frame_resized, f"FPS: {int(fps)}", (20, 40),
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
            
            results = self.model.track(
                frame_resized,
                conf=0.25,
                iou=0.6,
                classes=list(self.vehicle_classes.keys()),
                persist=True,
                verbose=False
            )

            # --- PERBAIKAN DI SINI ---
            for line in self.lines_info:
                (x1, y1), (x2, y2) = line['points']
                
                # 1. Ambil warna hex yang sudah disimpan (misal: "#FF0000" untuk merah)
                color_hex = line['color']
                
                # 2. Konversi warna hex ke format BGR yang dimengerti OpenCV
                #    (misal: "#FF0000" -> (0, 0, 255))
                color_bgr = tuple(int(color_hex.lstrip('#')[i:i+2], 16) for i in (4, 2, 0))
                
                # 3. Gunakan warna yang benar untuk menggambar garis
                cv2.line(frame_resized, (x1, y1), (x2, y2), color_bgr, 2)
            # --- AKHIR PERBAIKAN ---

            if results[0].boxes.id is not None:
                boxes = results[0].boxes.xyxy.cpu().numpy()
                class_ids = results[0].boxes.cls.cpu().numpy().astype(int)
                track_ids = results[0].boxes.id.cpu().numpy().astype(int)
                confidences = results[0].boxes.conf.cpu().numpy()

                for box, class_id, track_id, conf in zip(boxes, class_ids, track_ids, confidences):
                    if class_id not in self.vehicle_classes:
                        continue

                    class_name, hex_color = self.vehicle_classes[class_id]
                    color = tuple(int(hex_color.lstrip('#')[i:i+2], 16) for i in (4, 2, 0))
                    x1, y1, x2, y2 = map(int, box)
                    cx = (x1 + x2) // 2
                    cy = y2

                    self.update_counter(class_name, track_id, (cx, cy), conf)

                    trajectory = self.track_history.get(track_id, {}).get('positions', [])
                    if len(trajectory) > 1:
                        prev_pos = trajectory[-2]
                        current_pos = trajectory[-1]
                        arah = self.get_movement_direction(prev_pos, current_pos)
                        direction_text = arah.capitalize()
                        cv2.putText(frame_resized, direction_text, (x1, y1-10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)

                    cv2.rectangle(frame_resized, (x1, y1), (x2, y2), color, 2)
                    #cv2.putText(frame_resized, f"{conf:.2f}", (x1, y1-10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2) #Tampilkan confidence

                    cv2.putText(frame_resized, class_name, (x1, y1-30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)
                    cv2.circle(frame_resized, (cx, cy), 4, (0, 0, 255), -1)

            if self.frame_queue.qsize() < 10:
                frame_rgb = cv2.cvtColor(frame_resized, cv2.COLOR_BGR2RGB)
                self.frame_queue.put(frame_rgb)

        messagebox.showinfo("Selesai", "Video selesai diproses.")

        print("\n=== Hasil Perhitungan Kendaraan ===")
        print(f"Sumber: {self.video_path}")
        for line in self.lines_info:
            print(f"\nJalan {line['name']}:")
            for veh_name, count in line['counter'].items():
                print(f"  {veh_name}: {count}")
            print(f"  Total: {sum(line['counter'].values())}")

        print(f"Total Semua Kendaraan: {sum(self.total_vehicle_count.values())}")
        self.save_btn.config(state='normal')
        self.cap.release()

    def update_counter(self, class_name, track_id, position, conf):
        if track_id not in self.track_history:
            self.track_history[track_id] = {
                'positions': [],
                'counted_roads': set(),
                'confidences': {}
            }

        self.track_history[track_id]['positions'].append(position)

        if (class_name not in self.track_history[track_id]['confidences']):
            self.track_history[track_id]['confidences'][class_name] = []
        self.track_history[track_id]['confidences'][class_name].append(conf)

        trajectory = self.track_history[track_id]['positions']
        if len(trajectory) < 2:
            return

        prev_pos = trajectory[-2]
        current_pos = trajectory[-1]

        x_start, y_start = trajectory[0]
        x_end, y_end = trajectory[-1]
        dx, dy = x_end - x_start, y_end - y_start
        jarak = math.hypot(dx, dy)
                                                                            #         ↑ -90°
                                                                            #    ←          → 0°
                                                                            #         ↓ +90°
        arah = "diam"
        if jarak > 5:
            angle = math.degrees(math.atan2(-dy, dx))
            if 30 <= angle < 90:        # Rentang luas untuk bawah
                arah = "bawah"
            elif -45 <= angle <= 45:       # Gerakan kanan
                arah = "kanan"
            elif -110 < angle <= -90:        # Gerakan ke atas
                arah = "atas"
            elif angle > 70 or angle < -90:  # Gerakan kiri
                arah = "kiri"
        for line in self.lines_info:
            points = line['points']
            road_name = line['name']
            allowed_directions = line.get('allowed_directions', [])
            counter = line['counter']

            if self.line_intersection(points, (prev_pos, current_pos)):
                if arah in allowed_directions:
                    if (road_name, class_name) not in self.track_history[track_id]['counted_roads']:
                        counter[class_name] += 1
                        self.total_vehicle_count[class_name] += 1
                        self.track_history[track_id]['counted_roads'].add((road_name, class_name))

    def line_intersection(self, line1, line2):
        def ccw(A, B, C):
            return (C[1]-A[1]) * (B[0]-A[0]) > (B[1]-A[1]) * (C[0]-A[0])
        A, B = line1
        C, D = line2
        return ccw(A, C, D) != ccw(B, C, D) and ccw(A, B, C) != ccw(A, B, D)
    
    def get_movement_direction(self, prev, curr):
        dx = curr[0] - prev[0]
        dy = curr[1] - prev[1]
        if abs(dx) > abs(dy):
            return "kanan" if dx > 0 else "kiri"
        else:
            return "bawah" if dy > 0 else "atas"


    def calculate_average_confidence(self, road_name, class_name):
        total_conf = 0
        total_count = 0
        for track_data in self.track_history.values():
            key = (road_name, class_name)
            if key in track_data['confidences']:
                total_conf += sum(track_data['confidences'][key])
                total_count += len(track_data['confidences'][key])
        if total_count > 0:
            return (total_conf / total_count) * 100
        return None

    def update_gui(self):
        if not self.frame_queue.empty():
            frame = self.frame_queue.get()
            img = Image.fromarray(frame)
            self.tk_img = ImageTk.PhotoImage(image=img)
            self.canvas.create_image(0, 0, image=self.tk_img, anchor=tk.NW)

        with self.lock:
            for line in self.lines_info:
                road_name = line['name']
                total_per_road = 0
                for veh_name in line['counter']:
                    count = line['counter'][veh_name]
                    avg_conf = self.calculate_average_confidence(road_name, veh_name)
                    if avg_conf is not None:
                        text = f"{veh_name}: {count} ({avg_conf:.1f}%)"
                    else:
                        text = f"{veh_name}: {count}"
                    self.counter_labels[(road_name, veh_name)].config(text=text)
                    total_per_road += count
                self.counter_labels[(road_name, 'Total')].config(text=f"Total: {total_per_road}")

            total_count = sum(self.total_vehicle_count.values())
            self.total_count_label.config(text=f"Total Semua Kendaraan: {total_count}")

        if self.running:
            self.root.after(10, self.update_gui)

    def save_to_excel(self):
        if not self.lines_info:
            messagebox.showwarning("Peringatan", "Tidak ada data untuk disimpan.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Statistik Kendaraan"

        headers = ["Jalan", "Jenis Kendaraan", "Jumlah", "Rata-rata Prediksi (%)"]
        ws.append(headers)

        row_idx = 2  # Mulai dari baris ke-2 karena baris pertama untuk header

        for line in self.lines_info:
            road_name = line['name']
            start_row = row_idx
            for veh_name, count in line['counter'].items():
                avg_conf = self.calculate_average_confidence(road_name, veh_name)
                conf_text = f"{avg_conf:.2f}%" if avg_conf is not None else "-"
                ws.append([road_name, veh_name, count, conf_text])
                row_idx += 1

            # Tambah baris total untuk jalan ini
            total = sum(line['counter'].values())
            ws.append([road_name, "Total", total, "-"])
            row_idx += 1

            end_row = row_idx - 1
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                cell = ws.cell(row=start_row, column=1)
                cell.alignment = Alignment(vertical='center', horizontal='center')

        # Tambah total semua kendaraan
        ws.append(["SEMUA", "TOTAL SEMUA", sum(self.total_vehicle_count.values()), "-"])

        # Lebarkan kolom
        for i, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 22
        video_name = os.path.splitext(os.path.basename(self.video_path))[0]

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # Hasil nama file rapi
        filename = f"hasil_{video_name}_{timestamp}.xlsx"
        wb.save(filename)
        messagebox.showinfo("Berhasil", f"Hasil disimpan di: {filename}")
    def shutdown(self):
        self.running = False
        self.root.destroy()
        if hasattr(self, 'cap') and self.cap.isOpened():
            self.cap.release()

if __name__ == "__main__":
    root = tk.Tk()
    app = VehicleCounterApp(root)
    app.update_gui()
    root.mainloop()